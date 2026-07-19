import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models.contact import Contact, ContactProfession
from app.models.profession import Profession


def import_contacts_from_excel(file_stream, update_existing: bool, created_by_id: int):
    """Import contacts from a fixed-column spreadsheet: nombre, profesiones
    (comma-separated profession names, matched against the existing catalog -
    unmatched names are ignored, same convention as everywhere else in this
    app: never invent catalog entries on import).
    Returns (created, updated). Caller commits."""
    df = pd.read_excel(file_stream, dtype=str)
    df = df.where(pd.notna(df), None)
    df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

    all_professions = {p.name.lower(): p for p in Profession.query.all()}
    created = updated = 0

    for row in df.to_dict(orient='records'):
        nombre = str(row.get('nombre') or '').strip()
        if not nombre:
            continue

        existing = None
        if update_existing:
            existing = Contact.query.filter(db.func.lower(Contact.nombre) == nombre.lower()).first()

        if existing:
            contact = existing
            updated += 1
        else:
            contact = Contact(nombre=nombre, created_by_id=created_by_id)
            db.session.add(contact)
            created += 1

        contact.nombre = nombre
        db.session.flush()

        ContactProfession.query.filter_by(contact_id=contact.id).delete()
        for prof_name in str(row.get('profesiones') or '').split(','):
            prof = all_professions.get(prof_name.strip().lower())
            if prof:
                db.session.add(ContactProfession(contact_id=contact.id, profession_id=prof.id))

    return created, updated


def export_contacts_to_excel(contacts):
    """Build an Excel workbook (nombre, profesiones) from a list of Contact
    rows. Returns a BytesIO buffer ready to send as download."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contactos'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='center', wrap_text=True)

    thin = Side(style='thin', color='DEE2E6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Nombre', 'Profesiones']
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    ws.row_dimensions[1].height = 30

    for row_idx, contact in enumerate(contacts, start=2):
        values = [
            contact.nombre,
            ', '.join(cp.profession.name for cp in contact.professions),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
